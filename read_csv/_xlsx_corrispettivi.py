# -*- coding: utf-8 -*-
"""
Created on Sat Mar  8 18:12:34 2025

@author: Mirko
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os
import locale

# Imposta la lingua in italiano
locale.setlocale(locale.LC_TIME, 'it_IT.utf8')  

from constant import DOCS_TYPES


def xlsx_corrispettivi(self, dataframe, doc_type="CORRISPETTIVI"):
    
    if dataframe is None:
        return
    
    df = dataframe[0] 

    # Assicurati che la colonna 'Data e ora rilevazione' sia in formato datetime
    df['Data e ora rilevazione'] = pd.to_datetime(df['Data e ora rilevazione'], dayfirst=True)

    # Crea un nuovo workbook
    wb = Workbook()
    wb.remove(wb.active)  # Rimuove il foglio predefinito

    # Stili per la formattazione
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    border_style = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))
    center_align = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")
    
    
    # Ciclo per elaborare i singoli mesi dei corrispettivi
    for periodo, dfmth in df.groupby(df['Data e ora rilevazione'].dt.strftime('%Y-%m')):
        sheet_name = periodo.replace('-', '_')  # Sostituisce '-' con '_' per compatibilità Excel
        ws = wb.create_sheet(title=sheet_name)
        
        # Imposta la pagina per adattarsi alle colonne e formato A4 con margini normali
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.7
        ws.page_margins.right = 0.7
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3
        
        # Attiva l'adattamento alla pagina e imposta le dimensioni
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        # Ripeti le prime 5 righe in stampa
        ws.print_title_rows = "1:5"
        
        # Inserisci le informazioni del cliente
        ws['A1'] = "Cliente:"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].border = border_style
        
        ws['B1'] = self.cliente_folder
        ws['B1'].border = border_style
        
        ws['A2'] = "Periodo:"
        ws['A2'].font = header_font
        ws['A2'].fill = header_fill
        ws['A2'].border = border_style
        
        ws['B2'] = pd.to_datetime(periodo).strftime('%B %Y').capitalize()  # Ora periodo è una stringa compatibile
        ws['B2'].border = border_style
        
        # Imposta larghezza minima per le colonne A, B e C
        ws.column_dimensions['A'].width = max(ws.column_dimensions['A'].width, 13)
        ws.column_dimensions['B'].width = max(ws.column_dimensions['B'].width, 23)
        ws.column_dimensions['C'].width = max(ws.column_dimensions['C'].width, 23)
        
        # Scrivi i dati a partire dalla riga 5
        for r_idx, row in enumerate(dataframe_to_rows(dfmth, index=False, header=True), start=5):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.border = border_style
                if r_idx == 5:  # Formatta l'header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                
                # Allinea la colonna A, B e C al centro 
                if c_idx in [1, 2, 3]:
                    cell.alignment = center_align
                
                # Formatta la colonna C come data e ora
                if c_idx == 3:
                    cell.number_format = 'DD/MM/YYYY HH:MM:SS'
                
                # Formatta le colonne K-M come contabilità fino alla riga totale
                if c_idx in [11, 12, 13] and r_idx >= 6:
                    cell.number_format = '_-* #,##0.00 €_-;-* #,##0.00 €_-;_-* "-"?? €_-;_-@_-'
                    cell.alignment = center_align
        
        # Nascondi colonne specifiche (D-J)
        for col in range(4, 11):  # Colonne da D (4) a J (10)
            ws.column_dimensions[chr(64 + col)].hidden = True
        
        # Imposta larghezza colonne K, L, M a 15
        for col in ['K', 'L', 'M']:
            ws.column_dimensions[col].width = 15
        
        # Aggiungi riga di somma per le colonne K-M alla fine del foglio
        sum_row = ws.max_row + 2
        ws[f'C{sum_row}'] = "TOTALE"  # Etichetta somma in C
        ws[f'C{sum_row}'].font = Font(bold=True)
        ws[f'C{sum_row}'].border = border_style
        ws[f'C{sum_row}'].alignment = right_align
        for col in ['K', 'L', 'M']:
            ws[f'{col}{sum_row}'] = f"=SUM({col}6:{col}{sum_row-1})"
            ws[f'{col}{sum_row}'].font = Font(bold=True)
            ws[f'{col}{sum_row}'].border = border_style
            ws[f'{col}{sum_row}'].alignment = center_align
            ws[f'{col}{sum_row}'].number_format = '_-* #,##0.00 €_-;-* #,##0.00 €_-;_-* "-"?? €_-;_-@_-'
    
    
    # -------------- INIZIO BLOCCO RIEPILOGHI ----------------

    # Foglio di riepilogo mensile
    ws_mensile = wb.create_sheet(title="Riepilogo Mensile")

    # Raggruppa e calcola somme mensili
    df['Mese'] = df['Data e ora rilevazione'].dt.strftime('%Y-%m')
    riepilogo_mensile = df.groupby('Mese').agg({
        df.columns[10]: 'sum',  # Colonna K
        df.columns[11]: 'sum',  # Colonna L
        df.columns[12]: 'sum'   # Colonna M
    }).reset_index()

    # Scrivi intestazione riepilogo mensile
    intestazioni_mensili = ['Mese', 'Imponibile', 'IVA', 'Totale']
    ws_mensile.append(intestazioni_mensili)

    # Scrivi dati riepilogo mensile
    for row in riepilogo_mensile.itertuples(index=False):
        ws_mensile.append(row)

    # Aggiungi una riga vuota prima del totale mensile
    ws_mensile.append([])

    # Aggiungi riga totale mensile
    totale_mensile = ['TOTALE',
                      riepilogo_mensile[df.columns[10]].sum(),
                      riepilogo_mensile[df.columns[11]].sum(),
                      riepilogo_mensile[df.columns[12]].sum()]
    ws_mensile.append(totale_mensile)

    # Formatta intestazioni riepilogo mensile
    for cell in ws_mensile["1:1"]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = center_align

    # Formattazione celle riepilogo mensile
    for row in ws_mensile.iter_rows(min_row=2, min_col=1, max_col=4):
        for cell in row:
            cell.border = border_style
            cell.alignment = center_align
            if cell.column > 1:
                cell.number_format = '_-* #,##0.00 €_-;-* #,##0.00 €_-;_-* "-"?? €_-;_-@_-'

    # Formatta la riga totale mensile in grassetto
    for cell in ws_mensile[ws_mensile.max_row]:
        cell.font = Font(bold=True)

    # Imposta larghezza colonne riepilogo mensile
    for col, width in zip(['A', 'B', 'C', 'D'], [15, 15, 15, 15]):
        ws_mensile.column_dimensions[col].width = width

    # Foglio di riepilogo trimestrale
    ws_trimestrale = wb.create_sheet(title="Riepilogo Trimestrale")

    # Calcola trimestre e raggruppa
    df['Trimestre'] = df['Data e ora rilevazione'].dt.to_period('Q').astype(str)
    riepilogo_trimestrale = df.groupby('Trimestre').agg({
        df.columns[10]: 'sum',  # Colonna K
        df.columns[11]: 'sum',  # Colonna L
        df.columns[12]: 'sum'   # Colonna M
    }).reset_index()

    # Scrivi intestazione riepilogo trimestrale
    intestazioni_trimestrali = ['Trimestre', 'Imponibile', 'IVA', 'Totale']
    ws_trimestrale.append(intestazioni_trimestrali)

    # Scrivi dati riepilogo trimestrale
    for row in riepilogo_trimestrale.itertuples(index=False):
        ws_trimestrale.append(row)

    # Aggiungi una riga vuota prima del totale trimestrale
    ws_trimestrale.append([])

    # Aggiungi riga totale trimestrale
    totale_trimestrale = ['TOTALE',
                          riepilogo_trimestrale[df.columns[10]].sum(),
                          riepilogo_trimestrale[df.columns[11]].sum(),
                          riepilogo_trimestrale[df.columns[12]].sum()]
    ws_trimestrale.append(totale_trimestrale)

    # Formatta intestazioni riepilogo trimestrale
    for cell in ws_trimestrale["1:1"]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = center_align

    # Formattazione celle riepilogo trimestrale
    for row in ws_trimestrale.iter_rows(min_row=2, min_col=1, max_col=4):
        for cell in row:
            cell.border = border_style
            cell.alignment = center_align
            if cell.column > 1:
                cell.number_format = '_-* #,##0.00 €_-;-* #,##0.00 €_-;_-* "-"?? €_-;_-@_-'

    # Formatta la riga totale trimestrale in grassetto
    for cell in ws_trimestrale[ws_trimestrale.max_row]:
        cell.font = Font(bold=True)

    # Imposta larghezza colonne riepilogo trimestrale
    for col, width in zip(['A', 'B', 'C', 'D'], [15, 15, 15, 15]):
        ws_trimestrale.column_dimensions[col].width = width



        
    # SALVATAGGIO
    filename = self.make_filename_xlsx(doc_type)
    
    fullpath_xlsx = os.path.join(self.path_folder_iva, doc_type)
    
    if os.path.exists(fullpath_xlsx):
        print("⚠️ ATTENZIONE: File già creato!")
    else:
        wb.save(fullpath_xlsx)
        
    # Apri il file automaticamente
    os.system(f'"{filename}"')
    
       
    # Apri il file
    os.startfile(filename)
