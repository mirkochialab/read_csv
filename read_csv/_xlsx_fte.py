# -*- coding: utf-8 -*-
"""
Created on Sat Mar  8 18:12:34 2025

@author: Mirko
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os
import locale

# Imposta la lingua in italiano
locale.setlocale(locale.LC_TIME, 'it_IT.utf8')  

# from constant import DOCS_TYPES

def xlsx_fte(self, dataframe):
    
    doc_type = list(dataframe.keys())[0]
    df = dataframe[doc_type][0]
    
    
    if doc_type == "FTE_EMESSE":
        key_date = 'Data emissione'
    else:
        key_date = 'Data ricezione'
        
    # Assicurati che la colonna 'key_date' sia in formato datetime
    df[key_date] = pd.to_datetime(df[key_date], dayfirst=True)

    # Filtra in base alla periodicità del contribuente
    if self.periodicity_iva == "M":
        periodo_str = self.dt_chiusura_iva.strftime("%B %Y")  # Nome del mese
        df = df[df[key_date].dt.to_period("M") == self.dt_chiusura_iva.to_period("M")]
    else:  # Trimestrale
        trimestre = (self.dt_chiusura_iva.month - 1) // 3 + 1  # Numero del trimestre
        periodo_str = f"Trimestre {trimestre} {self.anno_iva}"
        df = df[df[key_date].dt.to_period("Q") == self.dt_chiusura_iva.to_period("Q")]
    
    # Seleziona le colonne che hanno nel testo la parola 'Data' o 'Periodo'
    date_cols = df.filter(regex='Data|Periodo').columns
    
    # Converti in formato "%d/%m/%Y" solo se il tipo è datetime
    for col in date_cols:
        df[col] = pd.to_datetime(df[col], errors='coerce', dayfirst=True).dt.strftime("%d/%m/%Y")
    
    # Crea un nuovo workbook
    wb = Workbook()
    wb.remove(wb.active)  # Rimuove il foglio predefinito
    ws = wb.create_sheet(title="Fatture Emesse")

    # Impostazioni pagina
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = '1:5'  # Ripete le prime 5 righe
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    
    # Stili per la formattazione
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    border_style = Border(left=Side(style='thin'),
                          right=Side(style='thin'),
                          top=Side(style='thin'),
                          bottom=Side(style='thin'))
    left_align = Alignment(horizontal="left")
    center_align = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")
    
    # Intestazione cliente
    ws["A1"].value = "Cliente"
    ws["A1"].font = header_font
    ws["A1"].border = border_style
    ws["A1"].alignment = left_align
    ws["B1"].value = self.cliente_folder
    ws["B1"].border = border_style
    
    # Intestazione protocolli
    ws["A2"].value = "Protocollo da"
    ws["A2"].font = header_font
    ws["A2"].border = border_style
    ws["A3"].value = "Protocollo aa"
    ws["A3"].font = header_font
    ws["A3"].border = border_style
    
    # Inserisci valori minimi e massimi nella colonna "Protocollo MIVA"
    ws["B2"].value = df["Protocollo MIVA"].min()
    ws["B2"].border = border_style
    ws["B2"].alignment = center_align
    ws["B3"].value = df["Protocollo MIVA"].max()
    ws["B3"].border = border_style
    ws["B3"].alignment = center_align
    
    # Imposta header del foglio con il tipo di documento e periodo
    ws.oddHeader.right.text = f"&R{doc_type} - {periodo_str.capitalize()}"
    ws.oddHeader.right.size = 12
    ws.oddHeader.right.font = "Calibri,Bold"
    
    # Stacco riga
    ws.append([])
    
    # Selezione colonne specifiche
    if doc_type == "FTE_EMESSE":
        
        df["Data ricezione"] = ""  # Aggiungi colonna vuota per uniformare la tabella
        
        colonne_da_mantenere = ["Protocollo MIVA", 
                                "Numero fattura / Documento", 
                                "Tipo documento", 
                                "Data emissione",
                                "Data ricezione", 
                                "Denominazione cliente", 
                                "Imponibile", "IVA", "TOTALE"]
        
    else:
        colonne_da_mantenere = ["Protocollo MIVA", 
                                "Numero fattura / Documento", 
                                "Tipo documento", 
                                "Data emissione", 
                                "Data ricezione", 
                                "Denominazione fornitore", 
                                "Imponibile", "IVA", "TOTALE"]
    
    df = df[colonne_da_mantenere]
    
    # Scrittura intestazione tabella dati
    row_idx = 5
    for col_idx, column_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=column_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_style
        cell.alignment = center_align
    
    # Scrittura dati
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=row_idx+1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border_style
            if c_idx == 6:  # Denominazione cliente
                cell.alignment = left_align
            else:
                cell.alignment = center_align
            if c_idx in [7, 8, 9] and isinstance(value, (int, float)):
                cell.number_format = '_-* #,##0.00 €_-;-* #,##0.00 €_-;_-* "-"?? €_-;_-@_-'
                cell.alignment = right_align
    
    # Imposta larghezza colonne
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 15.55
    ws.column_dimensions["E"].width = 15.55
    ws.column_dimensions["F"].width = 35
    
    for col in "GHI":
        ws.column_dimensions[col].width = 15
    
    # Aggiungi totale
    total_row = ws.max_row + 2
    ws[f"F{total_row}"].value = "TOTALE"
    ws[f"F{total_row}"].font = header_font
    ws[f"F{total_row}"].border = border_style
    for col in "GHI":
        ws[f"{col}{total_row}"].value = f"=SUM({col}6:{col}{total_row-2})"
        ws[f"{col}{total_row}"].number_format = '_-* #,##0.00 €_-;-* #,##0.00 €_-;_-* "-"?? €_-;_-@_-'
        ws[f"{col}{total_row}"].font = header_font
        ws[f"{col}{total_row}"].border = border_style
    
    # Nome del file
    filename = "{}_{}_{}_{}.xlsx".format(
        self.anno_iva,
        doc_type,
        self.cliente_folder,
        self.dt_chiusura_iva.strftime("%Y-%m")
    )
    
    # Salva il file
    wb.save(filename)
    print()
    print(f"File {filename} salvato con successo!")
    
    # Apri il file automaticamente
    os.system(f'"{filename}"')





